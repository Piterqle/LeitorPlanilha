from flask import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from datetime import datetime
import customtkinter as ctk

def insert_Str(liststr , insert_format):
    insert_format.sort(key=lambda x: x[0], reverse=True)
    
    for index, text in insert_format:
        liststr.insert(index, text)
    
    return "".join(liststr)

def addAluno(aluno, modalidade, data_marcada, data_experiencia, horario, contato, status, next):
    try:
        
        if([aluno, modalidade, data_marcada, data_experiencia, horario, contato, status].count("") > 0):
            print("Preencha todos os campos.")
            return
        
        if(datetime.strptime(data_experiencia, "%d/%m/%Y") < datetime.now()):
            print("A data marcada não pode ser anterior à data atual.")
            return
        
        if(contato[0] != "("):
            liststr = list(contato)
            insert_format = [(0, "("), (2, ")")]
            contato = insert_Str(liststr, insert_format)

        if(contato[11] != "-"):
            list(contato)
            insert_format = [(10, "-")]
            contato = insert_Str(liststr, insert_format)
        
        with open('Experimentais/caminho.json', 'r') as f:
            caminho = f.read()
            if caminho != "":
                caminho = json.loads(caminho)
                
                workbook = load_workbook(caminho["Planilha"])
                worksheet = workbook[modalidade.upper()]
                
                new_row = [aluno, data_marcada, data_experiencia, horario, contato, status]
                worksheet.append(new_row)
                
                if worksheet.tables:
                    table = list(worksheet.tables.values())[0]
                    table.ref = f"A1:F{worksheet.max_row}"

                workbook.save(caminho["Planilha"])
                next()
                
    except Exception as e:
        
        print("Erro ao adicionar aluno:", e)
        return
    
def editAluno(dados, id, entrys, buttons):
    if id.isalpha():
        print("ID não contem Letras")
        return
    
    # Preencher os campos com os dados    
    for i in range(len(entrys)):
        if isinstance(entrys[i], ctk.CTkEntry):
            entrys[i].delete(0, ctk.END)
            entrys[i].insert(0, dados[int(id)-1][i])
        elif isinstance(entrys[i], ctk.CTkComboBox):
            entrys[i].set(dados[int(id)-1][i])
    
    buttons[0].configure(fg_color="#6eff74", hover_color="#3ad63a", command=lambda: saveEdit(dados, id, entrys, buttons))
    
def saveEdit(dados, id, entrys, buttons):
    try:
        with open('Experimentais/caminho.json', 'r') as f:
            caminho = f.read()
            if caminho != "":
                caminho = json.loads(caminho)
                
                workbook = load_workbook(caminho["Planilha"])
                modalidade = entrys[1].get()
                worksheet = workbook[modalidade.upper()]
                
                new_row = [entry.get() if isinstance(entry, ctk.CTkEntry) else entry.get() for entry in entrys]
                
                for col_num, value in enumerate(new_row, start=1):
                    if col_num == 1: pass
                    worksheet.cell(row=int(id)+1, column=col_num, value=value)
                
                workbook.save(caminho["Planilha"])
                buttons[0].configure(fg_color="#d4b350", hover_color="#b38600", command=lambda: editAluno(dados, id, entrys, buttons))
                
    except Exception as e:
        
        print("Erro ao salvar edição do aluno:", e)
        return    