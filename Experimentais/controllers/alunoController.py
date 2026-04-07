from flask import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from datetime import datetime
import customtkinter as ctk


def insert_Str(liststr , insert_format):
    insert_format.sort(key=lambda x: x[0], reverse=True)
    
    for index, text in insert_format:
        liststr.insert(index, text)
    
    return "".join(liststr)


class alunoController():
    def __init__(self, entrys=None, buttons=None, next=None, dados=None, id=None):
        self.entrys = entrys
        self.buttons = buttons
        self.next = next
        self.dados = dados
        self.id = id

        pass
    
    def addAluno(self):
        try:
            [aluno, modalidade, data_marcada, data_experiencia, horario, contato, status] = [value.get() for value in self.entrys]
            if [aluno, modalidade, data_marcada, data_experiencia, horario, contato, status].count("") > 0:
                print("Preencha todos os campos.")
                return
            
            if datetime.strptime(data_experiencia, "%d/%m/%Y") < datetime.now():
                print("A data marcada não pode ser anterior à data atual.")
                return
            
            if contato[0] != "(" :
                liststr = list(contato)
                insert_format = [(0, "("), (2, ")")]
                contato = insert_Str(liststr, insert_format)

            if contato[11] != "-":
                list(contato)
                insert_format = [(10, "-")]
                contato = insert_Str(liststr, insert_format)

            with open('Experimentais/caminho.json', 'r') as f:
                caminho = f.read()
                if caminho != "":
                    caminho = json.loads(caminho)
                    
                    workbook = load_workbook(caminho["Planilha"])
                    worksheet = workbook[modalidade.upper()]
                    
                    new_row = [aluno, data_marcada, data_experiencia, horario, contato, status]
                    worksheet.append(new_row)
                    
                    if worksheet.tables:
                        table = list(worksheet.tables.values())[0]
                        table.ref = f"A1:F{worksheet.max_row}"

                    workbook.save(caminho["Planilha"])
                    self.next()
                    
        except Exception as e:
            
            print("Erro ao adicionar aluno:", e)
            return
        
    def editarAluno(self):
        if self.id.isalpha():
            print("ID não contem Letras")
            return
        
        
        # Preencher os campos com os dados  
        aluno = self.dados[int(self.id)-1]
        for i, value in enumerate(aluno.__dict__.values()):
            if isinstance(self.entrys[i], ctk.CTkEntry):
                self.entrys[i].delete(0, ctk.END)
                self.entrys[i].insert(0, value)
            elif isinstance(self.entrys[i], ctk.CTkComboBox):
                self.entrys[i].set(value)
        
        self.buttons[0].configure(hover_color="#2a632c", fg_color="#3ad63a", text="Salvar",command=lambda: self.salvarEdicao())
        self.buttons[1].configure(text="Cancelar", command=lambda: self.cancelarEdicao())

    def salvarEdicao(self):
        try:
            with open('Experimentais/caminho.json', 'r') as f:
                caminho = f.read()
                if caminho != "":
                    caminho = json.loads(caminho)
                    
                    workbook = load_workbook(caminho["Planilha"])
                    modalidade = self.entrys[1].get()
                    worksheet = workbook[modalidade.upper()]
                    
                    new_row = [entry.get() if isinstance(entry, ctk.CTkEntry) else entry.get() for entry in self.entrys]
                    
                    for col_num, value in enumerate(new_row, start=1):
                        if col_num == 1: pass
                        worksheet.cell(row=int(self.id)+1, column=col_num, value=value)
                    
                    workbook.save(caminho["Planilha"])
                    self.buttons[1].configure(text="Excluir", command=lambda: self.deletarAluno())
                    self.buttons[0].configure(fg_color="#d4b350", hover_color="#b38600", text="Editar",command=lambda: self.editarAluno())
                    
        except Exception as e:
            
            print("Erro ao salvar edição do aluno:", e)
            return 
        
    def cancelarEdicao(self):
        for entry in self.entrys:
            if isinstance(entry, ctk.CTkEntry):
                entry.delete(0, ctk.END)

        self.buttons[1].configure(text="Excluir", command=lambda: self.deletarAluno(self.dados, self.id))
        self.buttons[0].configure(text="Editar", fg_color="#d4b350", hover_color="#b38600")
        
        
    def deletarAluno(self):
        if self.id.isalpha():
            print("ID não contem Letras")
            return
        
        try:
            with open('Experimentais/caminho.json', 'r') as f:
                caminho = f.read()
                if caminho != "":
                    caminho = json.loads(caminho)
                    
                    workbook = load_workbook(caminho["Planilha"])
                    modalidade = self.dados[int(self.id)-1].modalidade
                    worksheet = workbook[modalidade.upper()]
                    
                    worksheet.delete_rows(int(self.dados[int(self.id)-1].row))
                    
                    if worksheet.tables:
                        table = list(worksheet.tables.values())[0]
                        table.ref = f"A1:F{worksheet.max_row}"
                    
                    workbook.save(caminho["Planilha"])
                self.next()
         
        except Exception as e:
            
            print("Erro ao excluir aluno:", e)
            return